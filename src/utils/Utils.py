from src.config import *
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import shutil
from datetime import datetime

class Utils:
    
    COR_HEADER  = PatternFill("solid", start_color="1F3864", end_color="1F3864")
    COR_ABA     = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
    COR_AUSENTE = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")   # 🔴 laranja
    COR_PARCIAL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")   # 🟡 amarelo
    COR_TOTAL   = PatternFill("solid", start_color="BDD7EE", end_color="BDD7EE")

    @staticmethod
    def _letra_para_col(letra):
        """Converte letra(s) de coluna Excel para número. Ex: A→1, G→7, AA→27."""
        resultado = 0
        for c in letra.upper():
            resultado = resultado * 26 + (ord(c) - ord('A') + 1)
        return resultado

    @staticmethod
    def _status_cobertura(cod, contas_bal_aba_keys, contas_na_aba):
        partes = cod.split('-')
        if len(partes) < 2:
            return 'faltando'

        pref = partes[0]
        cr   = partes[1]

        # 1. Sobe: verifica se algum ANCESTRAL está na aba
        pref_temp = pref
        while len(pref_temp) > 1:
            pref_temp = pref_temp[:-1]
            if f"{pref_temp}-{cr}" in contas_na_aba:
                return 'coberto'

        # 2. Desce: verifica se algum DESCENDENTE está na aba
        for outro in contas_bal_aba_keys:
            if outro == cod:
                continue
            partes_outro = outro.split('-')
            if len(partes_outro) < 2 or partes_outro[1] != cr:
                continue
            if partes_outro[0].startswith(pref) and partes_outro[0] != pref:
                if outro in contas_na_aba:
                    return 'parcial'

        return 'faltando'
    
    @staticmethod
    def consolidar_referencias(formulas_por_aba):
        referencias    = {}
        abas_com_conta = {}
        for aba, celulas in formulas_por_aba.items():
            for info in celulas.values():
                for cod in info['contas']:
                    referencias.setdefault(cod, set()).add(aba)
                    abas_com_conta.setdefault(aba, set()).add(cod)
        return referencias, abas_com_conta

    @staticmethod
    def extrair_cr(codigo):
        partes = codigo.split('-')
        return partes[-1].strip() if len(partes) >= 2 else 'SEM_CR'
    
    @staticmethod
    def _borda():
        s = Side(style='thin', color='CCCCCC')
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _fmt_header(ws):
        for cell in ws[1]:
            cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            cell.fill      = Utils.COR_HEADER  # 👈 Usando a cor definida na classe
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 18

    @staticmethod
    def _aplicar_estilo(ws, tipos, colunas_valor=None):
        mapa_cor = {
            'ausente': Utils.COR_AUSENTE, # 👈 Usando a cor definida na classe
            'parcial': Utils.COR_PARCIAL, # 👈 Usando a cor definida na classe
        }
        colunas_valor = colunas_valor or []
        for i, tipo in enumerate(tipos, start=2):
            cor = mapa_cor.get(tipo)
            for cell in ws[i]:
                if cor:
                    cell.fill = cor
                cell.border = Utils._borda()
                cell.font   = Font(name='Arial', size=9, color='000000')
                if cell.column in colunas_valor and cell.value is not None:
                    try:
                        float(cell.value)
                        cell.number_format = '#,##0.00;(#,##0.00);-'
                        cell.alignment     = Alignment(horizontal='right')
                    except (TypeError, ValueError):
                        pass

    @staticmethod
    def preparar_arquivo():
        arquivo_copia = ARQUIVO_ENTRADA
        arquivo_copia = os.path.abspath(arquivo_copia)

        if not os.path.exists(arquivo_copia):
            print(f"❌ Arquivo não encontrado: {arquivo_copia}")
            return None
        
        if not arquivo_copia.lower().endswith((".xlsb", ".xlsx")):
            print("❌ Arquivo inválido. Use .xlsb ou .xlsx")
            return None
        
        pasta_temp = os.path.abspath("data/temp")
        os.makedirs(pasta_temp, exist_ok=True)

        nome_base = os.path.basename(arquivo_copia)
        nome_sem_ext , ext = os.path.splitext(nome_base)

        timestamp = datetime.now().strftime(r"%Y%m%d_%H%M%S")
        nome_copia = f"{nome_sem_ext}_{timestamp}{ext}"

        caminho_destino = os.path.join(pasta_temp, nome_copia)

        try:
            shutil.copy2(arquivo_copia, caminho_destino)
        except Exception as e:
            print(f"❌ Erro ao copiar: {e}")
            return None

        print(f"📄 Trabalhando na cópia: {caminho_destino}")

        return caminho_destino