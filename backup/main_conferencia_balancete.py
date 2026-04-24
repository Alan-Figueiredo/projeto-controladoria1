"""
Comparador de Balancete por Aba — com Centro de Custo via PROCV
===============================================================
Para cada aba:
  1. Lê as contas referenciadas via =@CONTAS_BALANCETE("codigo")
  2. Lê os Centros de Resultado (CR) referenciados via PROCV
  3. Filtra o Balancete pelos CRs da aba
  4. Verifica hierarquia pai/filho e mostra apenas o que realmente falta
  5. Gera relatório com nome da conta, débito, crédito, saldo e status

Uso:
    python comparar_balancete.py

Requisitos:
    pip install pandas openpyxl pyxlsb pywin32
"""

import re
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import shutil
# ─────────────────────────────────────────────────────────────────
# CONFIGURAÇÕES — ajuste conforme seu arquivo
# ─────────────────────────────────────────────────────────────────
ARQUIVO_ENTRADA          = r"data/input/planilha-teste.xlsb"
ABA_BALANCETE            = "Balancete"
ARQUIVO_SAIDA            = r"data/output/resultado_comparacao.xlsx"

ABAS_IGNORAR_EXTRA       = {'MENU', 'Instruções', 'Aux', 'Config'}

COLUNA_CONTA_BALANCETE   = "Conta"
COLUNA_NOME_BALANCETE    = "Descrição"
COLUNA_CR_BALANCETE      = None          # extraído do código da conta
COLUNA_DEBITO_BALANCETE  = "Débitos"
COLUNA_CREDITO_BALANCETE = "Créditos"
COLUNA_VALOR_BALANCETE   = "SALDO MÊS"
# ─────────────────────────────────────────────────────────────────

REGEX_FORMULA         = re.compile(r'CONTAS_BALANCETE\s*\(\s*"([^"]+)"\s*\)', re.IGNORECASE)
REGEX_CODIGO          = re.compile(r'^\d[\d\-]*\d-\d{6}$')
REGEX_CR_VAL          = re.compile(r'C\.R\.\s*(\d{6})', re.IGNORECASE)
REGEX_CONTA_ANALITICA = re.compile(r'^\d{8,}-\d{6}$')   # mínimo 8 dígitos antes do hífen

XL_CELL_TYPE_FORMULAS = -4123


# ─── Auxiliares ──────────────────────────────────────────────────

def _letra_para_col(letra):
    """Converte letra(s) de coluna Excel para número. Ex: A→1, G→7, AA→27."""
    resultado = 0
    for c in letra.upper():
        resultado = resultado * 26 + (ord(c) - ord('A') + 1)
    return resultado


def _status_cobertura(cod, contas_bal_aba_keys, contas_na_aba):
    """
    Analisa a posição hierárquica da conta e retorna:

      'coberto'  → algum ancestral (pai, avô…) já está na aba
                   → conta não precisa aparecer no relatório

      'parcial'  → nenhum ancestral está na aba, mas algum descendente está
                   → pai está faltando mas filhos cobrem parte do valor
                   → aparece com aviso 🟡

      'faltando' → ninguém da família está na aba
                   → aparece como pendência 🔴

    Hierarquia detectada por prefixo:
      344001000003-001000  →  pai: 344001-001000
      Remove 1 dígito por vez do prefixo para subir.
      Adiciona 1 dígito por vez para descer (verifica filhos).
    """
    partes = cod.split('-')
    if len(partes) < 2:
        return 'faltando'

    pref = partes[0]
    cr   = partes[1]

    # 1. Sobe: verifica se algum ANCESTRAL está na aba
    pref_temp = pref
    while len(pref_temp) > 1:
        pref_temp = pref_temp[:-1]
        if f"{pref_temp}-{cr}" in contas_na_aba:
            return 'coberto'

    # 2. Desce: verifica se algum DESCENDENTE está na aba
    for outro in contas_bal_aba_keys:
        if outro == cod:
            continue
        partes_outro = outro.split('-')
        if len(partes_outro) < 2 or partes_outro[1] != cr:
            continue
        if partes_outro[0].startswith(pref) and partes_outro[0] != pref:
            if outro in contas_na_aba:
                return 'parcial'

    return 'faltando'


# ─── Leitura via win32com ─────────────────────────────────────────

def ler_formulas_e_crs_win32(wb_path, abas_ignorar):
    """
    Para cada aba retorna:
      formulas_por_aba : {aba: {(row,col): {formula, contas, value}}}
      crs_por_aba      : {aba: set(cr_codes)}
    """
    import win32com.client as win32

    abs_path = os.path.abspath(wb_path)
    print("  🔌 Abrindo arquivo com Excel (win32com)...")

    xl = win32.DispatchEx("Excel.Application")
    try:
        xl.Visible = False
    except Exception:
        pass
    try:
        xl.DisplayAlerts = False
    except Exception:
        pass

    formulas_por_aba = {}
    crs_por_aba      = {}

    try:
        wb = xl.Workbooks.Open(abs_path, ReadOnly=True, UpdateLinks=False)

        for sheet in wb.Sheets:
            nome = sheet.Name
            if nome in abas_ignorar:
                print(f"  ⏭️  Ignorando aba: '{nome}'")
                continue

            print(f"  📖 Lendo aba: '{nome}'")
            celulas_formula = {}
            crs_da_aba      = set()

            try:
                formula_range = sheet.UsedRange.SpecialCells(XL_CELL_TYPE_FORMULAS)
            except Exception:
                print(f"     ⏭️  Nenhuma fórmula encontrada. Pulando.")
                continue

            for area in formula_range.Areas:
                for cell in area.Cells:
                    try:
                        formula    = str(cell.Formula or '')
                        formula_br = str(cell.FormulaLocal or '')
                    except Exception:
                        continue

                    formula_up    = formula.upper()
                    formula_br_up = formula_br.upper()

                    # ── CONTAS_BALANCETE ──────────────────────────
                    if 'CONTAS_BALANCETE' in formula_up:
                        contas = REGEX_FORMULA.findall(formula)
                        if contas:
                            try:
                                val = float(cell.Value) if cell.Value is not None else 0.0
                            except (ValueError, TypeError):
                                val = 0.0
                            celulas_formula[(cell.Row, cell.Column)] = {
                                'formula': formula,
                                'contas':  contas,
                                'value':   val,
                            }

                    # ── PROCV / VLOOKUP → captura CR ─────────────
                    tem_procv = (any(x in formula_up    for x in ('PROCV', 'VLOOKUP')) or
                                 any(x in formula_br_up for x in ('PROCV', 'VLOOKUP')))

                    if tem_procv:
                        refs_chave = re.findall(
                            r'(?:PROCV|VLOOKUP)\s*\(\s*\$?([A-Z]+)\$?(\d+)\s*[;,]',
                            formula_br, re.IGNORECASE
                        )
                        for col_letra, row_num in refs_chave:
                            try:
                                val_celula = str(
                                    sheet.Cells(int(row_num),
                                                _letra_para_col(col_letra)).Value or ''
                                ).strip()
                                match = REGEX_CR_VAL.search(val_celula)
                                if match:
                                    cr_code = match.group(1)
                                    crs_da_aba.add(cr_code)
                                    print(f"     🏷️  CR capturado: {cr_code}  ('{val_celula}')")
                            except Exception:
                                pass

            if celulas_formula:
                formulas_por_aba[nome] = celulas_formula
                n_contas = sum(len(c['contas']) for c in celulas_formula.values())
                print(f"     → {len(celulas_formula)} células CONTAS_BALANCETE | "
                      f"{n_contas} referências | {len(crs_da_aba)} CRs via PROCV")
            else:
                print(f"     ⏭️  Nenhuma fórmula CONTAS_BALANCETE. Pulando.")

            if crs_da_aba:
                crs_por_aba[nome] = crs_da_aba

        wb.Close(SaveChanges=False)
    finally:
        xl.Quit()

    return formulas_por_aba, crs_por_aba


# ─── Leitura do Balancete ─────────────────────────────────────────

def _engine(path):
    return 'pyxlsb' if str(path).lower().endswith('.xlsb') else 'openpyxl'


def _br_to_float(s):
    try:
        return float(str(s).strip().replace('.', '').replace(',', '.'))
    except (ValueError, AttributeError):
        return None


def _detectar_col_conta(df):
    for col in df.columns:
        if str(col).strip().lower() == 'conta':
            return col
    for col in df.columns:
        try:
            amostra = df[col].dropna().astype(str).head(100)
            if sum(1 for x in amostra if REGEX_CODIGO.match(x.strip())) >= 3:
                return col
        except Exception:
            continue
    return None


def _detectar_col_nome(df, col_conta):
    candidatos = ['descrição', 'descricao', 'nome', 'name', 'description',
                  'histórico', 'historico', 'título', 'titulo']
    for col in df.columns:
        if col == col_conta:
            continue
        if str(col).strip().lower() in candidatos:
            return col
    for col in df.columns:
        if col == col_conta:
            continue
        amostra = df[col].dropna().astype(str).head(20)
        if amostra.str.len().mean() > 10:
            return col
    return None


def _detectar_col_cr(df):
    candidatos = ['cr', 'centro de resultado', 'centro resultado',
                  'centro de custo', 'centro custo', 'cc']
    for col in df.columns:
        if str(col).strip().lower() in candidatos:
            return col
    for col in df.columns:
        amostra = df[col].dropna().astype(str).head(50)
        if sum(1 for x in amostra if re.match(r'^\d{6}$', x.strip())) >= 3:
            return col
    return None


def _detectar_col_debito(df):
    candidatos = ['débito', 'debito', 'débitos', 'debitos', 'déb', 'deb', 'debit']
    for col in df.columns:
        if str(col).strip().lower() in candidatos:
            return col
    return None


def _detectar_col_credito(df):
    candidatos = ['crédito', 'credito', 'créditos', 'creditos', 'créd', 'cred', 'credit']
    for col in df.columns:
        if str(col).strip().lower() in candidatos:
            return col
    return None


def _detectar_col_valor(df, col_conta):
    prioridade = ['saldo mês', 'saldo mes', 'saldo atual', 'saldo_mes', 'saldo_atual']
    for col in df.columns:
        if str(col).strip().lower() in prioridade:
            return col
    for col in reversed(df.columns):
        if col == col_conta:
            continue
        amostra = df[col].dropna().astype(str).head(50)
        validos = sum(1 for x in amostra if _br_to_float(x) is not None)
        if validos >= 3:
            return col
    return None


def ler_balancete(wb_path, nome_aba):
    """Retorna {codigo_conta: {nome, cr, debito, credito, valor}}"""
    print(f"\n📊 Lendo balancete: '{nome_aba}'")

    # Detecta linha real do cabeçalho procurando por 'Conta'
    header_row = 0
    engine     = _engine(wb_path)
    for skip in range(10):
        df_test = pd.read_excel(wb_path, sheet_name=nome_aba, header=skip,
                                nrows=1, dtype=str, engine=engine)
        cols = [str(c).strip().lower() for c in df_test.columns]
        if any(c == 'conta' for c in cols):
            header_row = skip
            break

    df = pd.read_excel(wb_path, sheet_name=nome_aba, header=header_row,
                       dtype=str, engine=engine)
    for col in df.columns:
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(df[col])

    print(f"   → Colunas encontradas ({len(df.columns)}):")
    for c in df.columns:
        dtype   = df[c].dtype
        n_num   = pd.to_numeric(df[c], errors='coerce').notna().sum()
        amostra = df[c].dropna().astype(str).head(3).tolist()
        print(f"      [{dtype}] '{c}'  nums={n_num}  ex: {amostra}")

    col_conta   = COLUNA_CONTA_BALANCETE   or _detectar_col_conta(df)
    if col_conta is None:
        raise ValueError(f"Coluna de conta não encontrada em '{nome_aba}'. "
                         "Defina COLUNA_CONTA_BALANCETE.")

    col_nome    = COLUNA_NOME_BALANCETE    or _detectar_col_nome(df, col_conta)
    col_cr      = COLUNA_CR_BALANCETE      or _detectar_col_cr(df)
    col_debito  = COLUNA_DEBITO_BALANCETE  or _detectar_col_debito(df)
    col_credito = COLUNA_CREDITO_BALANCETE or _detectar_col_credito(df)
    col_valor   = COLUNA_VALOR_BALANCETE   or _detectar_col_valor(df, col_conta)

    if col_valor is None:
        raise ValueError("Coluna de saldo não encontrada. Defina COLUNA_VALOR_BALANCETE.")

    print(f"   → Conta: '{col_conta}' | Nome: '{col_nome}' | CR: '{col_cr}' | "
          f"Débito: '{col_debito}' | Crédito: '{col_credito}' | Saldo: '{col_valor}'")

    def _float(row, col):
        if col is None:
            return 0.0
        v = row.get(col)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 0.0
        try:
            return float(v)
        except (ValueError, TypeError):
            return 0.0

    mapa = {}
    for _, row in df.iterrows():
        cod = str(row[col_conta]).strip() if pd.notna(row[col_conta]) else ''
        if not REGEX_CODIGO.match(cod):
            continue

        nome_conta = ''
        if col_nome:
            v = row.get(col_nome)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                nome_conta = str(v).strip()

        cr_val = ''
        if col_cr:
            v = row.get(col_cr)
            if v is not None and not (isinstance(v, float) and pd.isna(v)):
                cr_val = str(v).strip()
        if not cr_val:
            partes = cod.split('-')
            cr_val = partes[-1].strip() if len(partes) >= 2 else ''

        debito  = _float(row, col_debito)
        credito = _float(row, col_credito)
        valor   = _float(row, col_valor)

        if cod in mapa:
            mapa[cod]['debito']  += debito
            mapa[cod]['credito'] += credito
            mapa[cod]['valor']   += valor
        else:
            mapa[cod] = {
                'nome':    nome_conta,
                'cr':      cr_val,
                'debito':  debito,
                'credito': credito,
                'valor':   valor,
            }

    print(f"   → {len(mapa)} contas no Balancete")
    return mapa


# ─── Consolidação de referências ──────────────────────────────────

def consolidar_referencias(formulas_por_aba):
    referencias    = {}
    abas_com_conta = {}
    for aba, celulas in formulas_por_aba.items():
        for info in celulas.values():
            for cod in info['contas']:
                referencias.setdefault(cod, set()).add(aba)
                abas_com_conta.setdefault(aba, set()).add(cod)
    return referencias, abas_com_conta


def extrair_cr(codigo):
    partes = codigo.split('-')
    return partes[-1].strip() if len(partes) >= 2 else 'SEM_CR'


# ─── Formatação ───────────────────────────────────────────────────

def _cor(h):
    return PatternFill("solid", start_color=h, end_color=h)

COR_HEADER  = _cor("1F3864")
COR_ABA     = _cor("2E75B6")
COR_AUSENTE = _cor("FCE4D6")   # 🔴 laranja — falta completamente
COR_PARCIAL = _cor("FFF2CC")   # 🟡 amarelo — filhos cobrem, mas pai não está
COR_TOTAL   = _cor("BDD7EE")


def _borda():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)


def _fmt_header(ws):
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill      = COR_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 18


def _auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        w = max((len(str(c.value or '')) for c in col), default=min_w)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 3, max_w)


def _aplicar_estilo(ws, tipos, colunas_valor=None):
    mapa_cor = {
        'ausente': COR_AUSENTE,
        'parcial': COR_PARCIAL,
    }
    colunas_valor = colunas_valor or []
    for i, tipo in enumerate(tipos, start=2):
        cor = mapa_cor.get(tipo)
        for cell in ws[i]:
            if cor:
                cell.fill = cor
            cell.border = _borda()
            cell.font   = Font(name='Arial', size=9, color='000000')
            if cell.column in colunas_valor and cell.value is not None:
                try:
                    float(cell.value)
                    cell.number_format = '#,##0.00;(#,##0.00);-'
                    cell.alignment     = Alignment(horizontal='right')
                except (TypeError, ValueError):
                    pass


# ─── Geração do Relatório ─────────────────────────────────────────

def gerar_relatorio(mapa_balancete, referencias, abas_com_conta,
                    formulas_por_aba, crs_por_aba, arquivo_saida):

    # 1. Inverte a lógica: Descobre quais Abas cuidam de cada CR
    abas_por_cr = {}
    for aba, crs in crs_por_aba.items():
        for cr in crs:
            abas_por_cr.setdefault(cr, set()).add(aba)

    # Garante que abas sem PROCV (mas com contas) entrem no mapa
    for aba, contas in abas_com_conta.items():
        if aba not in crs_por_aba or not crs_por_aba[aba]:
            for c in contas:
                cr = extrair_cr(c)
                abas_por_cr.setdefault(cr, set()).add(aba)

    # ─────────────────────────────────────────────────────────────
    # NOVO: Trava de Limpeza (Remove FIN/ADM das abas específicas)
    for cr in list(abas_por_cr.keys()):
        # Pega todas as abas vinculadas ao CR, mas ignora ADM e FIN
        abas_especificas = {a for a in abas_por_cr[cr] if a not in ("ADM e Dir", "FIN e N.O.")}
        
        # Se houver alguma aba específica (ex: Acessórios), ela ganha exclusividade
        if abas_especificas:
            abas_por_cr[cr] = abas_especificas
    # ─────────────────────────────────────────────────────────────

    # 2. Agrupa todas as contas encontradas nas planilhas por CR
    contas_nas_planilhas_por_cr = {}
    for conta in referencias.keys():
        cr = extrair_cr(conta)
        contas_nas_planilhas_por_cr.setdefault(cr, set()).add(conta)

    COLUNAS  = ['CR', 'Abas Relacionadas', 'Conta', 'Nome da Conta', 
                'Débito', 'Crédito', 'Saldo', 'Status', 'Observação']
    COLS_NUM = [5, 6, 7] # Índices das colunas de valores

    # Protege contra arquivo aberto
    if os.path.exists(arquivo_saida):
        try:
            os.rename(arquivo_saida, arquivo_saida)
        except OSError:
            base, ext    = os.path.splitext(arquivo_saida)
            arquivo_saida = f"{base}_novo{ext}"
            print(f"  ⚠️  Arquivo em uso — salvando como '{arquivo_saida}'")

    with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
        linhas, tipos = [], []
        todos_crs = sorted(abas_por_cr.keys())

        for cr in todos_crs:
            abas_do_cr = " e ".join(sorted(abas_por_cr[cr]))
            contas_planilha = contas_nas_planilhas_por_cr.get(cr, set())

            # Puxa do Balancete apenas as contas que pertencem a este CR
            contas_bal_cr = {
                cod: info for cod, info in mapa_balancete.items() if info['cr'] == cr
            }
            contas_bal_cr_keys = set(contas_bal_cr.keys())

            pendencias = []
            for cod, info in contas_bal_cr.items():
                if cod in contas_planilha:
                    continue   # Conta já lançada nas abas -> OK

                status = _status_cobertura(cod, contas_bal_cr_keys, contas_planilha)

                if status == 'coberto':
                    continue   # Conta pai já cobre

                pendencias.append((cod, info, status))

            if not pendencias:
                continue   # CR sem pendências, pula para o próximo

            pendencias.sort(key=lambda x: x[0])

            for cod, info, status in pendencias:
                if status == 'parcial':
                    st_cell = '🟡 Parcial'
                    obs     = 'Filhos cobertos, mas conta pai não referenciada'
                else:
                    st_cell = '🔴 Falta na Aba'
                    obs     = f'Conta não encontrada na(s) aba(s): {abas_do_cr}'

                linhas.append({
                    'CR':                cr,
                    'Abas Relacionadas': abas_do_cr,
                    'Conta':             cod,
                    'Nome da Conta':     info['nome'],
                    'Débito':            info['debito']  if info['debito']  != 0 else None,
                    'Crédito':           info['credito'] if info['credito'] != 0 else None,
                    'Saldo':             info['valor']   if info['valor']   != 0 else None,
                    'Status':            st_cell,
                    'Observação':        obs,
                })
                tipos.append('ausente' if status == 'faltando' else 'parcial')

            

        df = pd.DataFrame(linhas, columns=COLUNAS)
        df.to_excel(writer, sheet_name='Contas Faltando', index=False)

        ws = writer.sheets['Contas Faltando']
        _fmt_header(ws)

        larguras = {'A': 10, 'B': 25, 'C': 25, 'D': 42,
                    'E': 16, 'F': 16, 'G': 16, 'H': 18, 'I': 60}
        for col_letra, w in larguras.items():
            ws.column_dimensions[col_letra].width = w

        _aplicar_estilo(ws, tipos, colunas_valor=COLS_NUM)

    print(f"\n✅ Relatório gerado: {arquivo_saida}")


def preparar_arquivo():
    arquivo_copia = ARQUIVO_ENTRADA
    arquivo_copia = os.path.abspath(arquivo_copia)

    if not os.path.exists(arquivo_copia):
        print(f"❌ Arquivo não encontrado: {arquivo_copia}")
    
    if not arquivo_copia.lower().endswith((".xlsb", ".xlsx")):
        print("❌ Arquivo inválido. Use .xlsb ou .xlsx")
        return None
    
    pasta_temp = os.path.abspath("data/temp")
    os.makedirs(pasta_temp,exist_ok=True)

    nome_base = os.path.basename(arquivo_copia)
    nome_sem_ext , ext = os.path.splitext(nome_base)

    timesamp = datetime.now().strftime(r"%Y%m%d_%H%M%S")
    nome_copia = f"{nome_sem_ext}_{timesamp}{ext}"

    caminho_destino = os.path.join(pasta_temp,nome_copia)

    try:
        shutil.copy2(arquivo_copia,caminho_destino)
    except Exception as e:
        print(f"❌ Erro ao copiar: {e}")
        return None

    print(f"📄 Trabalhando na cópia: {caminho_destino}")

    return caminho_destino




# ─── Entry Point ─────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Comparador de Balancete focado em Centro de Custo (CR)")
    print("=" * 60)

    ARQUIVO_ENTRADA = preparar_arquivo()

    if not ARQUIVO_ENTRADA:
        return

    
    if not os.path.exists(ARQUIVO_ENTRADA):
        print(f"\n❌ Arquivo '{ARQUIVO_ENTRADA}' não encontrado.")
        return

    abas_ignorar = ABAS_IGNORAR_EXTRA | {ABA_BALANCETE}

    # 1. Extrai fórmulas CONTAS_BALANCETE + CRs via PROCV
    formulas_por_aba, crs_por_aba = ler_formulas_e_crs_win32(ARQUIVO_ENTRADA, abas_ignorar)

    if not formulas_por_aba:
        print("\n❌ Nenhuma fórmula CONTAS_BALANCETE encontrada.")
        return

    # 2. Consolida referências
    referencias, abas_com_conta = consolidar_referencias(formulas_por_aba)
    print(f"\n✔  Contas únicas nas abas:  {len(referencias)}")
    print(f"✔  Abas com CRs via PROCV:  {len(crs_por_aba)}")

    # 3. Lê o Balancete
    mapa_balancete = ler_balancete(ARQUIVO_ENTRADA, ABA_BALANCETE)

    # 4. Gera relatório (agora por CR)
    gerar_relatorio(mapa_balancete, referencias, abas_com_conta,
                    formulas_por_aba, crs_por_aba, ARQUIVO_SAIDA)

    # 5. Resumo no terminal (agora focado em CR)
    print("\n📊 Resumo de Pendências por Centro de Custo (CR):")
    
    shutil.rmtree("data/temp")
    # Reconstrói o mapa de CRs para o resumo visual
    abas_por_cr = {}
    for aba, crs in crs_por_aba.items():
        for cr in crs:
            abas_por_cr.setdefault(cr, set()).add(aba)

    # ─────────────────────────────────────────────────────────────
    # NOVO: Mesma Trava de Limpeza no resumo de tela
    for cr in list(abas_por_cr.keys()):
        abas_especificas = {a for a in abas_por_cr[cr] if a not in ("ADM e Dir", "FIN e N.O.")}
        if abas_especificas:
            abas_por_cr[cr] = abas_especificas
    # ─────────────────────────────────────────────────────────────

    for cr in sorted(abas_por_cr.keys()):
        abas_do_cr = ", ".join(sorted(abas_por_cr[cr]))
        contas_planilha = {c for c in referencias.keys() if extrair_cr(c) == cr}
        contas_bal_cr = {cod for cod, info in mapa_balancete.items() if info['cr'] == cr}
        
        faltando = sum(
            1 for cod in contas_bal_cr
            if cod not in contas_planilha
            and _status_cobertura(cod, contas_bal_cr, contas_planilha) != 'coberto'
        )
        
        status = '✅' if faltando == 0 else f'⚠️  {faltando} pendência(s)'
        print(f"   {status}  [CR {cr}]  Planilhas: {abas_do_cr}")

if __name__ == "__main__":
    main()